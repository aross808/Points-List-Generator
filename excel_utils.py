from __future__ import annotations

import re
from typing import Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Font, Border
from openpyxl.utils import get_column_letter

from excel_styles import (
    FONT_MAIN, FONT_TITLE, FONT_SECTION,
    FILL_SECTION, ALIGN_LEFT,
    BORDER_THIN, THIN, MED,
    MAX_WIDTH, EXTRA_PAD,
)

def ensure_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(title=name)

def sanitize_sheet_title(title: str) -> str:
    """
    Excel sheet constraints:
      - max length 31
      - cannot contain: : \\ / ? * [ ]
    """
    bad = r'[:\\/?*\[\]]'
    t = re.sub(bad, " ", str(title))
    t = " ".join(t.split()).strip()
    if not t:
        t = "Sheet"
    return t[:31]

def unique_sheet_title(wb: Workbook, base_title: str) -> str:
    """Return a unique sheet title within this workbook (handles duplicates + 31-char limit)."""
    base = sanitize_sheet_title(base_title)
    if base not in wb.sheetnames:
        return base

    for i in range(2, 1000):
        suffix = f" ({i})"
        trimmed = base[: max(0, 31 - len(suffix))].rstrip()
        candidate = f"{trimmed}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate

    return sanitize_sheet_title(f"{base[:25]}_{len(wb.sheetnames)}")

def apply_sheet_basics(ws):
    ws.sheet_format.defaultRowHeight = 16

def autosize_columns(ws, *, extra_padding: int = EXTRA_PAD, max_width: int = MAX_WIDTH, skip_cols: Optional[Set[int]] = None):
    skip_cols = skip_cols or set()
    for col_cells in ws.columns:
        if not col_cells:
            continue
        col_idx = col_cells[0].column
        if col_idx in skip_cols:
            continue
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max_len + extra_padding, max_width)

def write_cell(ws, addr: str, value, *, font: Optional[Font] = None, fill=None, align=None, border: Optional[Border] = None):
    c = ws[addr]
    c.value = value
    c.font = font or FONT_MAIN
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border

def write_sheet_title(ws, title: str):
    write_cell(ws, "A1", title, font=FONT_TITLE, align=ALIGN_LEFT)
    ws.row_dimensions[1].height = 26

def write_section_bar(ws, row: int, text: str, start_col: int = 1, end_col: int = 8):
    for c in range(start_col, end_col + 1):
        addr = f"{get_column_letter(c)}{row}"
        write_cell(ws, addr, "" if c != start_col else text, font=FONT_SECTION, fill=FILL_SECTION, border=BORDER_THIN)
    ws.row_dimensions[row].height = 18

def find_table_end_col(ws, header_row: int, start_col: int = 1) -> int:
    c = start_col
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v in (None, ""):
            return c - 1
        c += 1

def outer_border(row: int, col: int, top: int, left: int, bottom: int, right: int) -> Border:
    left_side = MED if col == left else THIN
    right_side = MED if col == right else THIN
    top_side = MED if row == top else THIN
    bot_side = MED if row == bottom else THIN
    return Border(left=left_side, right=right_side, top=top_side, bottom=bot_side)

def apply_thick_separator(ws, row: int, left_col: int, right_col: int):
    """Adds a medium bottom border across a row (used as a section separator)."""
    from openpyxl.styles import Border
    for c in range(left_col, right_col + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border or BORDER_THIN
        cell.border = Border(
            left=b.left,
            right=b.right,
            top=b.top,
            bottom=MED,
        )
