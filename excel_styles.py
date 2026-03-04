from __future__ import annotations

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ----------------------------
# Theme / style constants
# ----------------------------
FONT_MAIN = Font(name="Calibri", size=11)
FONT_TITLE = Font(name="Calibri", size=18, bold=True)
FONT_SECTION = Font(name="Calibri", size=12, bold=True)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # dark blue
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")  # light blue/gray
FILL_ZEBRA = PatternFill("solid", fgColor="F7F7F7")    # very light gray

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

THIN = Side(style="thin", color="9E9E9E")
MED = Side(style="medium", color="000000")

BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=THIN, right=THIN, top=THIN, bottom=MED)

MAX_WIDTH = 70
EXTRA_PAD = 3
