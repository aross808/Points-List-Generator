from __future__ import annotations

import os
from typing import Optional, List, Dict

import pandas as pd
from openpyxl import Workbook

from excel_renderers import (
    render_point_selection_sheet,
    render_global_caiso_sheet,
    render_global_substation_sheet,
    render_meter_sheets,
    autosize_saved_workbook,
)


def write_workbook(
    output_path: str,
    *,
    point_selection_df: Optional[pd.DataFrame],
    caiso_df: Optional[pd.DataFrame],
    substation_df: Optional[pd.DataFrame],
    meters: Optional[List[Dict]] = None,
    autosize: bool = True,
) -> str:
    """
    Generates the output Excel workbook.

    FIXED METER LOGIC:
      - Meter tabs are derived ONLY from CAISO points
      - Rows are selected by CAISO DNP index
      - Tag prefix becomes <METERNAME>. (spaces → underscores)
    """
    out = os.path.abspath(output_path)

    wb = Workbook()
    wb.active.title = "Sheet"

    ps_grid_cols = None

    # --------------------------------------------------
    # POINT SELECTION
    # --------------------------------------------------
    if point_selection_df is not None and not point_selection_df.empty:
        ps_grid_cols = render_point_selection_sheet(
            wb, point_selection_df, meters=meters
        )

    # --------------------------------------------------
    # GLOBAL CAISO
    # --------------------------------------------------
    if caiso_df is not None and not caiso_df.empty:
        render_global_caiso_sheet(wb, caiso_df)

    # --------------------------------------------------
    # GLOBAL SUBSTATION
    # --------------------------------------------------
    if substation_df is not None and not substation_df.empty:
        render_global_substation_sheet(wb, substation_df)

    # --------------------------------------------------
    # PER-METER TABS (CAISO ONLY)
    # --------------------------------------------------
    if meters and caiso_df is not None and not caiso_df.empty:
        render_meter_sheets(
            wb,
            caiso_df=caiso_df,     # 🔒 force CAISO only
            sub_df=None,           # 🚫 ignore substation
            meters=meters,
        )

    # --------------------------------------------------
    # CLEANUP + SAVE
    # --------------------------------------------------
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    wb.save(out)

    if autosize:
        autosize_saved_workbook(out, ps_grid_cols=ps_grid_cols)

    return out
