from __future__ import annotations

from datetime import datetime
from typing import Optional, Tuple, List, Any, Set, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_styles import *
from excel_utils import (
    ensure_sheet,
    unique_sheet_title,
    apply_sheet_basics,
    autosize_columns,
    outer_border,
    apply_thick_separator,
    find_table_end_col,
)

# ============================================================
# Alignment configuration
# ============================================================

CENTER_COLUMNS = {
    "IP Address",
    "0 Means",
    "1 Means",
    "Analog Unit in RIG",
    "Scaling Applied in SCADA",
    "Group/Variation",
}

# ============================================================
# Shared writing helpers
# ============================================================

def write_cell(ws, addr: str, value, font=None, fill=None, align=None, border=None):
    c = ws[addr]
    c.value = value
    c.font = font or FONT_MAIN
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border


def write_sheet_title(ws, title: str):
    write_cell(ws, "A1", title, font=FONT_TITLE, align=ALIGN_LEFT)
    ws.row_dimensions[1].height = 26


def write_section_bar(ws, row: int, text: str, start_col: int = 1, end_col: int = 8):
    for c in range(start_col, end_col + 1):
        addr = f"{get_column_letter(c)}{row}"
        write_cell(
            ws,
            addr,
            "" if c != start_col else text,
            font=FONT_SECTION,
            fill=FILL_SECTION,
            border=BORDER_THIN,
        )
    ws.row_dimensions[row].height = 18


def write_point_selection_blocks(ws):
    write_section_bar(ws, 3, "Project Info", 1, 8)

    write_cell(ws, "B4", "Site Name", font=FONT_BOLD)
    write_cell(ws, "B5", "Generated", font=FONT_BOLD)
    write_cell(ws, "C5", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), font=FONT_MAIN)

    write_cell(ws, "E4", "RIG IP", font=FONT_BOLD)
    write_cell(ws, "E5", "Default Gateway", font=FONT_BOLD)
    write_cell(ws, "E6", "Subnet Mask", font=FONT_BOLD)
    write_cell(ws, "E7", "RIG Common Name", font=FONT_BOLD)


def write_caiso_blocks(ws):
    """
    Restores the metadata block (labels) that exists on the original CAISO tab.
    Values are intentionally left blank for user entry; we only print the labels.
    """
    write_section_bar(ws, 3, "Connection Info", 1, 8)
    labels = [
        ("A4", "CAISO RIG"),
        ("A5", "Protocol"),
        ("A6", "RTAC Device/Map"),
        ("A7", "Transport/Protocol"),
        ("A8", "Tunneling Method"),
        ("A9", "DNP Client Address"),
        ("A10", "Client IP Address"),
        ("A11", "DNP Server Address"),
        ("A12", "TCP Server Port"),
        ("A14", "CAISO Server"),
    ]
    for addr, txt in labels:
        write_cell(ws, addr, txt, font=FONT_BOLD)


def write_substation_blocks(ws):
    """
    Restores the metadata block (labels) that exists on the original Substation tab.
    Values are intentionally left blank for user entry; we only print the labels.
    """
    write_section_bar(ws, 3, "Connection Info", 1, 8)
    labels = [
        ("A4", "Substation RTU to CAISO RIG"),
        ("A5", "Protocol"),
        ("A6", "RTAC Device/Map"),
        ("A7", "Transport/Protocol"),
        ("A8", "Tunneling Method"),
        ("A9", "DNP Client Address"),
        ("A10", "Client IP Address"),
        ("A11", "DNP Server Address"),
        ("A12", "TCP Server Port"),
        ("A14", "CAISO Server"),
    ]
    for addr, txt in labels:
        write_cell(ws, addr, txt, font=FONT_BOLD)


def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> Tuple[int, int]:
    if df is None or df.empty:
        return start_row, start_col

    r0 = start_row
    c0 = start_col

    # header
    for j, col_name in enumerate(df.columns, start=c0):
        cell = ws.cell(row=r0, column=j, value=str(col_name))
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_HEADER

    # data
    for i, row in enumerate(df.itertuples(index=False, name=None), start=r0 + 1):
        is_zebra = ((i - (r0 + 1)) % 2) == 1
        for j, val in enumerate(row, start=c0):
            col_name = str(df.columns[j - c0])

            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONT_MAIN
            cell.border = BORDER_THIN

            # ✅ Center specific columns, otherwise left align
            cell.alignment = ALIGN_CENTER if col_name in CENTER_COLUMNS else ALIGN_LEFT

            if is_zebra:
                cell.fill = FILL_ZEBRA

    end_row = r0 + len(df)
    end_col = c0 + len(df.columns) - 1
    return end_row, end_col


def add_table(ws, top_row: int, left_col: int, bottom_row: int, right_col: int, table_name: str):
    ref = f"{get_column_letter(left_col)}{top_row}:{get_column_letter(right_col)}{bottom_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.auto_filter.ref = ref


def set_index_grid_widths(ws, start_col: int):
    for offset in (0, 2, 4, 6):
        ws.column_dimensions[get_column_letter(start_col + offset)].width = 5.0
    for offset in (1, 3, 5):
        ws.column_dimensions[get_column_letter(start_col + offset)].width = 2.0


def write_index_block(
    ws,
    *,
    title: str,
    start_col: int,
    header_row: int,
    data_start_row: int,
    row_count: int,
    kind_per_row: List[Any],
    value_per_row: List[Any],
):
    kinds = ["AI", "AO", "DI", "DO"]
    kind_to_offset = {"AI": 0, "AO": 2, "DI": 4, "DO": 6}

    top = header_row
    left = start_col
    right = start_col + 6
    bottom = data_start_row + row_count - 1

    # Title row
    for c in range(left, right + 1):
        cell = ws.cell(row=header_row, column=c, value="")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = outer_border(header_row, c, top, left, bottom, right)

    ws.cell(row=header_row, column=left, value=title)
    ws.merge_cells(start_row=header_row, start_column=left, end_row=header_row, end_column=right)

    # Subheader row
    sub_r = header_row + 1
    for c in range(left, right + 1):
        cell = ws.cell(row=sub_r, column=c, value="")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = outer_border(sub_r, c, top, left, bottom, right)

    for k in kinds:
        c = left + kind_to_offset[k]
        ws.cell(row=sub_r, column=c, value=k)

    # Body rows
    for r in range(row_count):
        excel_r = data_start_row + r
        row_kind = str(kind_per_row[r]).strip().upper()
        row_val = value_per_row[r]

        for c in range(left, right + 1):
            cell = ws.cell(row=excel_r, column=c, value="")
            cell.alignment = ALIGN_CENTER
            cell.border = outer_border(excel_r, c, top, left, bottom, right)

        if row_kind in kind_to_offset:
            c = left + kind_to_offset[row_kind]
            ws.cell(row=excel_r, column=c, value=row_val)


def write_df_sections_by_kind(
    ws,
    df: pd.DataFrame,
    *,
    start_row: int,
    kind_col: str = "__kind",
    first_col_name: str = "Point Description",
    order: Tuple[str, ...] = ("AI", "AO", "DI", "DO"),
) -> Tuple[int, int]:
    """
    Writes df in sections AI/AO/DI/DO, repeating headers each section.
    NOTE: __dnp is FILTER-ONLY and is always dropped before rendering.
    """
    if df is None or df.empty:
        return start_row, 1

    df = df.drop(columns=["__dnp"], errors="ignore")

    if kind_col not in df.columns:
        return write_df(ws, df, start_row=start_row, start_col=1)

    cur_row = start_row
    last_end_r, last_end_c = start_row, 1

    kinds_present: List[str] = []
    for k in order:
        part = df[df[kind_col].astype(str).str.upper() == k]
        if not part.empty:
            kinds_present.append(k)

    for idx, k in enumerate(kinds_present):
        part = df[df[kind_col].astype(str).str.upper() == k].copy()
        part = part.drop(columns=[kind_col])

        if first_col_name in part.columns:
            part = part.rename(columns={first_col_name: f"{first_col_name} ({k})"})

        last_end_r, last_end_c = write_df(ws, part, start_row=cur_row, start_col=1)

        if idx != len(kinds_present) - 1:
            apply_thick_separator(ws, row=last_end_r, left_col=1, right_col=last_end_c)

        cur_row = last_end_r + 1

    return last_end_r, last_end_c


# ============================================================
# Sheet renderers
# ============================================================

def render_point_selection_sheet(
    wb: Workbook,
    point_selection_df: pd.DataFrame,
    *,
    meters: Optional[List[dict]] = None,
) -> Tuple[int, int, List[int]]:
    ws = ensure_sheet(wb, "POINT SELECTION")
    apply_sheet_basics(ws)
    write_sheet_title(ws, "POINT SELECTION")
    write_point_selection_blocks(ws)

    meta_cols = ["caiso_kind", "caiso_index", "sub_kind", "sub_index"]
    missing = [c for c in meta_cols if c not in point_selection_df.columns]
    if missing:
        raise ValueError(f"POINT SELECTION missing meta columns: {missing}")

    meta = point_selection_df[meta_cols].copy()
    table_df = point_selection_df.drop(columns=meta_cols)

    # ------------------------------------------------------------
    # Row kind inference (AI/AO/DI/DO) for index grids
    # ------------------------------------------------------------
    def _norm_kind(x) -> str:
        s = str(x).strip().upper()
        return s if s in ("AI", "AO", "DI", "DO") else ""

    def _kind_from_point_type(v) -> str:
        s = str(v).strip().upper()
        if s in ("ANALOG INPUT", "AI"):
            return "AI"
        if s in ("ANALOG OUTPUT", "AO"):
            return "AO"
        if s in ("DIGITAL INPUT", "DI"):
            return "DI"
        if s in ("DIGITAL OUTPUT", "DO"):
            return "DO"
        return ""

    fallback_col = None
    for cand in ("Point Type", "PointType", "Type", "__kind", "kind"):
        if cand in table_df.columns:
            fallback_col = cand
            break

    row_kind: List[str] = []
    for i in range(len(meta)):
        k = _norm_kind(meta["caiso_kind"].iloc[i]) if "caiso_kind" in meta.columns else ""
        if not k:
            k = _norm_kind(meta["sub_kind"].iloc[i]) if "sub_kind" in meta.columns else ""
        if not k and fallback_col:
            if fallback_col == "Point Type":
                k = _kind_from_point_type(table_df[fallback_col].iloc[i])
            else:
                k = _norm_kind(table_df[fallback_col].iloc[i])
        row_kind.append(k)

    start_row = 10
    end_r, end_c = write_df(ws, table_df, start_row=start_row, start_col=1)

    add_table(ws, top_row=start_row, left_col=1, bottom_row=end_r, right_col=end_c, table_name="PointSelection")
    ws.freeze_panes = f"A{start_row+1}"

    row_count = len(table_df)
    table_data_start_row = start_row + 1

    caiso_grid_col = end_c + 1
    sub_grid_col = caiso_grid_col + 7

    set_index_grid_widths(ws, caiso_grid_col)
    set_index_grid_widths(ws, sub_grid_col)

    write_index_block(
        ws,
        title="CAISO (RIG IS SERVER)",
        start_col=caiso_grid_col,
        header_row=start_row - 2,
        data_start_row=table_data_start_row,
        row_count=row_count,
        kind_per_row=row_kind,
        value_per_row=meta["caiso_index"].tolist(),
    )

    write_index_block(
        ws,
        title="SUBSTATION",
        start_col=sub_grid_col,
        header_row=start_row - 2,
        data_start_row=table_data_start_row,
        row_count=row_count,
        kind_per_row=row_kind,
        value_per_row=meta["sub_index"].tolist(),
    )

    # Meter ledger blocks: headers only (no indices printed)
    meters = meters or []
    meter_cols: List[int] = []
    meter_start_col = sub_grid_col + 7

    for mm in meters:
        label = str(mm.get("label", "")).strip()
        dnp_set = set(mm.get("dnp", set()) or set())
        if not label or not dnp_set:
            continue

        meter_cols.append(meter_start_col)
        set_index_grid_widths(ws, meter_start_col)

        meter_values = [""] * row_count

        write_index_block(
            ws,
            title=label,
            start_col=meter_start_col,
            header_row=start_row - 2,
            data_start_row=table_data_start_row,
            row_count=row_count,
            kind_per_row=row_kind,
            value_per_row=meter_values,
        )

        meter_start_col += 7

    return (caiso_grid_col, sub_grid_col, meter_cols)


def render_global_caiso_sheet(wb: Workbook, caiso_df: pd.DataFrame):
    ws = ensure_sheet(wb, "CAISO (RIG IS SERVER)")
    apply_sheet_basics(ws)
    write_sheet_title(ws, "CAISO (RIG IS SERVER)")
    write_caiso_blocks(ws)

    start_row = 16
    write_df_sections_by_kind(ws, caiso_df, start_row=start_row)
    ws.freeze_panes = f"A{start_row}"


def render_global_substation_sheet(wb: Workbook, sub_df: pd.DataFrame):
    ws = ensure_sheet(wb, "SUBSTATION")
    apply_sheet_basics(ws)
    write_sheet_title(ws, "SUBSTATION")
    write_substation_blocks(ws)

    start_row = 16
    write_df_sections_by_kind(ws, sub_df, start_row=start_row)
    ws.freeze_panes = f"A{start_row}"


def render_meter_sheets(
    wb: Workbook,
    caiso_df: Optional[pd.DataFrame],
    sub_df: Optional[pd.DataFrame],
    meters: Optional[List[Dict]],
):
    """
    One sheet per meter label.

    FIXED BEHAVIOR:
      - Meter tabs are derived ONLY from CAISO df (ignore sub_df)
      - Rows selected by CAISO __dnp ∈ meter set
      - RTAC-RIG Tag Name = CAISO RIG Tag Name, with prefix replaced:
            CAISO_DNP.<kind>_####_<abbr>  ->  <METERNAME>.<kind>_####_<abbr>
        where <METERNAME> has spaces replaced with underscores
      - Column mapping is normalized so values actually populate.
    """
    meters = meters or []
    base_df = caiso_df  # 🔒 CAISO only

    def df_filter_by_dnp(df: Optional[pd.DataFrame], dnp_set: Set[int]) -> Optional[pd.DataFrame]:
        if df is None or df.empty:
            return None
        if "__dnp" in df.columns:
            return df[df["__dnp"].isin(dnp_set)]
        if "caiso_index" in df.columns:
            return df[df["caiso_index"].isin(dnp_set)]
        return None

    # Meter template columns (what the sheet expects)
    COLS = [
        "Point Description",
        "RTAC-RIG Tag Name",
        "Analog Unit in SUBSTATION RTU",
        "Scaling Applied in SUBSTATION RTU",
        "Group/Variation",
        "0 Means",
        "1 Means",
    ]

    def _meterize_prefix(tag: str, meter_label: str) -> str:
        s = str(tag or "").strip()
        if not s:
            return ""
        meter_prefix = meter_label.strip().replace(" ", "_")
        # Only rewrite CAISO tags; leave anything else alone
        if s.startswith("CAISO_DNP."):
            return meter_prefix + "." + s[len("CAISO_DNP."):]
        return s

    def _prepare(df: pd.DataFrame, meter_label: str) -> pd.DataFrame:
        """
        Normalize CAISO dataframe columns -> meter template columns.
        """
        df = df.copy()

        # Drop filter-only column if present
        df = df.drop(columns=["__dnp"], errors="ignore")

        # Normalize names (CAISO -> meter template)
        rename_map = {
            "RIG Tag Name": "RTAC-RIG Tag Name",
            "Analog Unit to CAISO": "Analog Unit in SUBSTATION RTU",
            "Scaling Applied in RIG": "Scaling Applied in SUBSTATION RTU",
        }
        df = df.rename(columns=rename_map)

        # Ensure all expected columns exist
        for c in COLS:
            if c not in df.columns:
                df[c] = ""

        # Apply meter prefix rewrite to tag name column
        df["RTAC-RIG Tag Name"] = df["RTAC-RIG Tag Name"].map(lambda x: _meterize_prefix(x, meter_label))

        return df[COLS]

    def write_header_only(ws, columns: list[str], row: int, start_col: int = 1) -> tuple[int, int]:
        for j, col_name in enumerate(columns, start=start_col):
            cell = ws.cell(row=row, column=j, value=str(col_name))
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_HEADER
        end_col = start_col + len(columns) - 1
        return row, end_col

    def _write_section(ws, *, kind: str, df_kind: Optional[pd.DataFrame], meter_label: str, start_row: int) -> int:

        cols = list(COLS)
        cols[0] = f"Point Description ({kind})"

        # ----- Empty section -----
        if df_kind is None or df_kind.empty:

            # Write header row
            end_r, end_c = write_header_only(ws, cols, row=start_row, start_col=1)

            # Add one blank spacer row
            spacer_row = end_r + 1
            for c in range(1, end_c + 1):
                cell = ws.cell(row=spacer_row, column=c, value="")
                cell.border = BORDER_THIN
                cell.fill = None

            return spacer_row + 1

        # ----- Normal section -----
        sec_df = _prepare(df_kind, meter_label)
        sec_df = sec_df.rename(columns={"Point Description": f"Point Description ({kind})"})

        end_r, end_c = write_df(ws, sec_df, start_row=start_row, start_col=1)

        apply_thick_separator(ws, row=end_r, left_col=1, right_col=end_c)

        return end_r + 1

    for mm in meters:
        label = str(mm.get("label", "")).strip()
        dnp_set = set(mm.get("dnp", set()) or set())
        if not label or not dnp_set:
            continue

        df_m = df_filter_by_dnp(base_df, dnp_set)

        sheet_name = unique_sheet_title(wb, label)
        ws = ensure_sheet(wb, sheet_name)
        apply_sheet_basics(ws)

        write_sheet_title(ws, sheet_name)
        write_substation_blocks(ws)

        cur_row = 16

        if df_m is not None and not df_m.empty and "__kind" in df_m.columns:
            def part(k: str) -> pd.DataFrame:
                return df_m[df_m["__kind"].astype(str).str.upper() == k].copy()

            cur_row = _write_section(ws, kind="AI", df_kind=part("AI"), meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="AO", df_kind=part("AO"), meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="DI", df_kind=part("DI"), meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="DO", df_kind=part("DO"), meter_label=label, start_row=cur_row)
        else:
            cur_row = _write_section(ws, kind="AI", df_kind=df_m, meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="AO", df_kind=None, meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="DI", df_kind=None, meter_label=label, start_row=cur_row)
            cur_row = _write_section(ws, kind="DO", df_kind=None, meter_label=label, start_row=cur_row)

        ws.freeze_panes = "A16"



def autosize_saved_workbook(out_path: str, *, ps_grid_cols: Optional[Tuple[int, int, List[int]]]):
    """
    Autosize all sheets, but keep index grids on POINT SELECTION at fixed widths.
    """
    from openpyxl import load_workbook

    wb2 = load_workbook(out_path)

    if "POINT SELECTION" in wb2.sheetnames:
        ws = wb2["POINT SELECTION"]

        if ps_grid_cols is None:
            end_c_table = find_table_end_col(ws, header_row=10, start_col=1)
            caiso_grid_col = end_c_table + 1
            sub_grid_col = caiso_grid_col + 7
            meter_cols: List[int] = []
        else:
            caiso_grid_col, sub_grid_col, meter_cols = ps_grid_cols

        skip = set(range(caiso_grid_col, caiso_grid_col + 7)) | set(range(sub_grid_col, sub_grid_col + 7))
        for mc in (meter_cols or []):
            skip |= set(range(mc, mc + 7))

        autosize_columns(ws, skip_cols=skip)

        set_index_grid_widths(ws, caiso_grid_col)
        set_index_grid_widths(ws, sub_grid_col)
        for mc in (meter_cols or []):
            set_index_grid_widths(ws, mc)

    for name in wb2.sheetnames:
        if name == "POINT SELECTION":
            continue
        autosize_columns(wb2[name])

    wb2.save(out_path)
